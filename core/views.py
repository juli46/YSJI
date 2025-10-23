# --------- IMPORTACIONES ---------
# Manejo de datos JSON
import json
# Para crear decoradores
from functools import wraps
# Renderizado de vistas, redirección y obtención segura de objetos
from django.shortcuts import render, redirect, get_object_or_404
# Mensajes temporales al usuario
from django.contrib import messages
# Encriptar y verificar contraseñas
from django.contrib.auth.hashers import make_password, check_password
# Restringir vistas a método POST
from django.views.decorators.http import require_POST
# Respuestas JSON para AJAX o APIs
from django.http import JsonResponse
# Generar URLs por nombre
from django.urls import reverse
# Consultas complejas
from django.db.models import Q
from django.core.exceptions import FieldError
# Manejo de fechas y horas con zona horaria
from django.utils import timezone

# Modelos
from .models import (
    Usuario, Producto, Categoria, Pedido, ImagenProducto, Marca, Proveedor, Mensaje
)
# Formularios
from .forms import (RegistroForm, ProductoForm, CategoriaForm, UsuarioForm, MarcaForm, ProveedorForm, MensajeForm
)

# ...existing code...
import json
import os
import logging
import requests
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.db import transaction

logger = logging.getLogger(__name__)

# Usa variables de entorno en producción
PAYPAL_CLIENT = os.environ.get('PAYPAL_CLIENT_ID', 'AfFmoNGhHVg4hXOHy6wtl10KQZNu9qmKPy5oNMw3EFtvJfQtvb_JdyZyv2W2fB30HMqDv1qym_FBIZ6v')
PAYPAL_SECRET = os.environ.get('PAYPAL_SECRET', 'EDIzN1By-VddpXTSsQw7yIWLYQd9mKYOL2ZWYXBomFlXnblS29eQW4AYnLlBPP9fB6kEdkeU0x0rCoQU')
PAYPAL_API_BASE = 'https://api-m.sandbox.paypal.com'

def _get_paypal_token():
    resp = requests.post(
        f"{PAYPAL_API_BASE}/v1/oauth2/token",
        auth=(PAYPAL_CLIENT, PAYPAL_SECRET),
        data={'grant_type': 'client_credentials'},
        timeout=10
    )
    resp.raise_for_status()
    return resp.json().get('access_token')

@require_POST
def paypal_confirm(request):
    """
    Recibe JSON { orderID, carrito, payer? }.
    Verifica la orden en PayPal y crea un Pedido local si está COMPLETED.
    Responde JSON claro para depuración.
    """
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception as e:
        logger.exception("paypal_confirm: JSON inválido")
        return HttpResponseBadRequest(json.dumps({'ok': False, 'error': 'invalid_json', 'detail': str(e)}), content_type='application/json')

    order_id = payload.get('orderID')
    carrito = payload.get('carrito', [])
    logger.info("paypal_confirm: recibida orderID=%s carrito_items=%s", order_id, len(carrito))

    if not order_id:
        logger.warning("paypal_confirm: falta orderID")
        return JsonResponse({'ok': False, 'error': 'missing_orderID'}, status=400)

    # Obtener token
    try:
        token = _get_paypal_token()
    except Exception as e:
        logger.exception("paypal_confirm: fallo al obtener token PayPal")
        return JsonResponse({'ok': False, 'error': 'token_error', 'detail': str(e)}, status=500)

    # Obtener información de la orden
    try:
        r = requests.get(f"{PAYPAL_API_BASE}/v2/checkout/orders/{order_id}",
                         headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
                         timeout=10)
        r.raise_for_status()
        order_info = r.json()
        logger.debug("paypal_confirm: order_info=%s", order_info)
    except requests.HTTPError as he:
        logger.exception("paypal_confirm: PayPal returned HTTP error")
        body = getattr(he.response, 'text', str(he))
        return JsonResponse({'ok': False, 'error': 'paypal_http_error', 'detail': body}, status=502)
    except Exception as e:
        logger.exception("paypal_confirm: error llamando a PayPal")
        return JsonResponse({'ok': False, 'error': 'paypal_fetch_error', 'detail': str(e)}, status=500)

    # Revisar estado: puede estar en CREATED / APPROVED / COMPLETED
    status = (order_info.get('status') or '').upper()
    logger.info("paypal_confirm: order %s status=%s", order_id, status)

    # Alternativa: comprobar captures dentro de purchase_units -> payments -> captures
    captures = []
    for pu in order_info.get('purchase_units', []):
        payments = pu.get('payments') or {}
        captures += payments.get('captures', []) if payments else []

    # Si hay captures comprobar alguno COMPLETED
    capture_completed = any((c.get('status') or '').upper() == 'COMPLETED' for c in captures)

    if status in ('COMPLETED', 'APPROVED') or capture_completed:
        # Guardar pedido localmente (ajusta según tu modelo Pedido)
        try:
            # obtener usuario desde sesión (esta app usa sesión personalizada)
            usuario = None
            if request.session.get('usuario_id'):
                usuario = Usuario.objects.filter(id=request.session['usuario_id']).first()
            if not usuario:
                logger.warning("paypal_confirm: usuario no autenticado en sesión, no se puede crear pedido")
                return JsonResponse({'ok': False, 'error': 'not_authenticated'}, status=403)

            # carrito puede venir como lista o string JSON; normalizar a lista de dicts
            if isinstance(carrito, str):
                try:
                    carrito = json.loads(carrito)
                except Exception:
                    carrito = []

            productos_sanitizados = []
            total_pedido = 0.0
            for it in (carrito or []):
                if not isinstance(it, dict):
                    continue
                precio = 0.0
                cantidad = 1
                try:
                    precio = float(it.get('precio') or it.get('price') or 0)
                except Exception:
                    precio = 0.0
                try:
                    cantidad = int(it.get('cantidad') or it.get('qty') or 1)
                except Exception:
                    cantidad = 1
                nombre = it.get('nombre') or it.get('name') or ''
                codigo = it.get('codigo') or it.get('sku') or ''
                imagen = it.get('imagen') or it.get('image') or ''
                productos_sanitizados.append({
                    'nombre': nombre,
                    'precio': precio,
                    'cantidad': cantidad,
                    'imagen': imagen,
                    'codigo': codigo
                })
                total_pedido += precio * cantidad

            # Crear pedido (productos como lista -> JSONField)
            pedido = Pedido.objects.create(
                usuario=usuario,
                productos=productos_sanitizados,
                estado='Pendiente'
            )

            # Si la tabla DB tiene columna `valor_total`, intentar actualizarla (encapsulado)
            try:
                Pedido.objects.filter(id=pedido.id).update(valor_total=total_pedido)
            except Exception:
                # si la columna no existe o falla, no romper el flujo
                logger.debug("paypal_confirm: no se actualizó valor_total (columna ausente o error)")

            logger.info("paypal_confirm: Pedido creado id=%s order_id=%s", pedido.id, order_id)
            return JsonResponse({'ok': True, 'status': status, 'order_info': order_info, 'pedido_id': pedido.id})
        except Exception as e:
            logger.exception("paypal_confirm: fallo guardando Pedido")
            return JsonResponse({'ok': False, 'error': 'db_error', 'detail': str(e)}, status=500)
    else:
        logger.warning("paypal_confirm: orden no completada status=%s", status)
        return JsonResponse({'ok': False, 'status': status, 'order_info': order_info}, status=400)

# --------- CONTEXT PROCESSORS / UTILIDADES ---------

def get_usuario_admin(request):
    """Retorna el usuario admin logueado desde la sesión."""
    usuario_admin = None
    if request.session.get('usuario_id'):
        try:
            usuario_admin = Usuario.objects.get(id=request.session['usuario_id'])
        except Usuario.DoesNotExist:
            pass
    return {'usuario_admin': usuario_admin}

def sesion_usuario(request):
    """Retorna si hay una sesión activa."""
    return {'sesion_activa': 'usuario_id' in request.session}

# --------- DECORADORES ---------

def login_required(view_func):
    """Verifica que el usuario esté autenticado."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if 'usuario_id' not in request.session:
            return redirect('login')
        return view_func(request, *args, **kwargs) 
    return wrapper


# --------- AUTH / SESIÓN ---------
def login_register_view(request):
    form = RegistroForm()
    if request.session.get('usuario_id'):
        return redirect('cuenta')

    if request.method == 'POST':
        # LOGIN AJAX
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' and 'login' in request.POST:
            correo = request.POST.get('correo', '').strip()
            contraseña = request.POST.get('contraseña', '')
            if not correo or not contraseña:
                return JsonResponse({'success': False, 'message': 'Todos los campos son obligatorios.'})
            try:
                usuario = Usuario.objects.get(correo=correo)
                if check_password(contraseña, usuario.contraseña):
                    request.session['usuario_id'] = usuario.id
                    request.session['usuario_nombre'] = usuario.nombre
                    request.session['usuario_rol'] = usuario.rol
                    redirect_url = reverse('dashboard') if usuario.rol == 'admin' else reverse('index')
                    return JsonResponse({
                        'success': True,
                        'message': f'¡Bienvenido {usuario.nombre}!',
                        'redirect_url': redirect_url
                    })
                else:
                    return JsonResponse({'success': False, 'message': 'Credenciales inválidas.'})
            except Usuario.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Correo no registrado.'})

        # REGISTRO (POST normal)
        elif 'register' in request.POST:
            form = RegistroForm(request.POST)
            if form.is_valid():
                nombre = form.cleaned_data['nombre']
                correo = form.cleaned_data['correo']
                contraseña = make_password(form.cleaned_data['contraseña'])
                if Usuario.objects.filter(correo=correo).exists():
                    form.add_error('correo', 'Este correo ya está registrado.')
                else:
                    Usuario.objects.create(
                        nombre=nombre,
                        correo=correo,
                        contraseña=contraseña,
                        rol=getattr(Usuario, 'USUARIO', 'usuario')  
                    )
                    messages.success(request, f'¡{nombre}, tu cuenta ha sido creada exitosamente!')
                    return redirect('login')
            else:
                messages.error(request, "Por favor corrige los errores en el formulario.")
            return render(request, 'login.html', {'form': form})

    return render(request, 'login.html', {'form': form})

def logout_view(request):
    """Cerrar sesión del usuario."""
    request.session.flush()
    return redirect('index')

# --------- PERFIL DE USUARIO ---------
@login_required
def cuenta_view(request):
    """Página de cuenta del usuario."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id'])
    pedidos = Pedido.objects.filter(usuario_id=usuario.id).order_by('-fecha')
    return render(request, 'cuenta.html', {'usuario': usuario, 'pedidos': pedidos})

@login_required
def editar_perfil_view(request):
    """Permite al usuario editar su perfil."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id'])
    if request.method == 'POST':
        usuario.nombre = request.POST.get('nombre')
        usuario.correo = request.POST.get('correo')
        if request.FILES.get('imagen_perfil'):
            usuario.imagen_perfil = request.FILES['imagen_perfil']
        contrasena_actual = request.POST.get('contraseña_actual')
        nueva_contrasena = request.POST.get('contraseña_nueva')
        confirmar_contrasena = request.POST.get('contraseña_confirmar')
        if nueva_contrasena:
            if not contrasena_actual or not check_password(contrasena_actual, usuario.contraseña):
                messages.error(request, 'La contraseña actual es incorrecta.')
                return redirect('cuenta')
            if nueva_contrasena != confirmar_contrasena:
                messages.error(request, 'La nueva contraseña y la confirmación no coinciden.')
                return redirect('cuenta')
            usuario.contraseña = make_password(nueva_contrasena)
        usuario.save()
        messages.success(request, 'Perfil actualizado correctamente.')
    return redirect('cuenta')

# --------- PÁGINAS PÚBLICAS ---------
from django.shortcuts import render, get_object_or_404
from .models import Usuario, Blog, Producto

def index_view(request):
    """Página de inicio."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id']) if 'usuario_id' in request.session else None
    
    blogs = Blog.objects.all().order_by('-id')[:3]
    
    # Solo 4 productos (puedes cambiarlos después a más vendidos cuando tengas ventas)
    productos_destacados = Producto.objects.all()[:4]

    return render(request, 'index.html', {
        'usuario': usuario,
        'sesion_activa': usuario is not None,
        'blogs': blogs,
        'productos_destacados': productos_destacados
    })

def SobreNosotros_view(request):
    """Página Sobre Nosotros."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id']) if 'usuario_id' in request.session else None
    return render(request, 'SobreNosotros.html', {'usuario': usuario, 'sesion_activa': usuario is not None})

def Contactenos_view(request):
    """Página Contáctenos."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id']) if 'usuario_id' in request.session else None
    return render(request, 'Contactenos.html', {'usuario': usuario, 'sesion_activa': usuario is not None})

# --------- CATÁLOGO Y DETALLE DE PRODUCTO ---------

def Catalogo_view(request):
    """Vista del catálogo de productos con filtros por categoría y marca."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id']) if 'usuario_id' in request.session else None
    categoria_filtro = request.GET.get('categoria')
    marca_filtro = request.GET.get('marca')
    productos = Producto.objects.all()
    if categoria_filtro and categoria_filtro != 'Todas':
        # Si el parámetro es numérico, lo tratamos como id de la categoría (FK).
        # Si no, lo tratamos como nombre y filtramos por nombre de categoría (FK).
        # Intentamos mantener compatibilidad con el campo legado 'categoria_producto' si existe;
        # si no existe, capturamos FieldError y filtramos solo por la FK.
        if categoria_filtro.isdigit():
            productos = productos.filter(categoria__id=int(categoria_filtro))
        else:
            try:
                productos = productos.filter(
                    Q(categoria__nombre__iexact=categoria_filtro) | Q(categoria_producto__iexact=categoria_filtro)
                )
            except FieldError:
                productos = productos.filter(categoria__nombre__iexact=categoria_filtro)
    if marca_filtro and marca_filtro != 'Todas':
        productos = productos.filter(marca__nombre__iexact=marca_filtro)
    
    categorias_menu = ['Camisetas', 'Abrigos', 'Sacos', 'Chaquetas', 'Pantalones', 'Jeans', 'Faldas', 'Vestidos', 'Accesorios', 'Bolsos']
    conteos = {cat: 0 for cat in categorias_menu}
    for producto in Producto.objects.all():
        # Preferir la FK 'categoria' si existe, sino usar el campo legado
        if getattr(producto, 'categoria', None):
            cat_display = producto.categoria.nombre
        else:
            cat_display = getattr(producto, 'categoria_producto', '') or ''
        if cat_display in conteos:
            conteos[cat_display] += 1
    categorias_conteo = [(cat, conteos[cat]) for cat in categorias_menu]
    marcas = Marca.objects.all().order_by('nombre')
    return render(request, 'Catalogo.html', {
        'productos': productos,
        'categorias_conteo': categorias_conteo,
        'marcas': marcas,
        'usuario': usuario,
        'sesion_activa': usuario is not None,
    })

def producto_detalle_dashboard(request, pk):
    """Detalle de producto para el dashboard admin."""
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'producto_detalle_dashboard.html', {'producto': producto})

def producto_detalle_publico(request, pk):
    """Detalle de producto público (cliente)."""
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'producto_detalle_publico.html', {'producto': producto})

@login_required
def producto_detalle(request, producto_id):
    """Detalle de producto para usuarios logueados."""
    producto = get_object_or_404(Producto, pk=producto_id)
    return render(request, 'producto_detalle.html', {'producto': producto})

# --------- VISTAS DE PEDIDO ---------

@login_required
def cancelar_pedido_view(request, pedido_id):
    """Cancelar pedido si está pendiente."""
    usuario = get_object_or_404(Usuario, id=request.session['usuario_id'])
    pedido = get_object_or_404(Pedido, id=pedido_id, usuario_id=usuario.id)
    if pedido.estado != "Pendiente":
        messages.error(request, "Solo puedes cancelar pedidos pendientes.")
        return redirect('cuenta')
    if request.method == "POST":
        pedido.estado = "Cancelado"
        pedido.save()
        messages.success(request, "Solicitud de cancelación enviada.")
        return redirect('cuenta')
    return render(request, "cancelar_pedido_confirm.html", {"pedido": pedido})
# --------- DASHBOARD ADMIN ---------

@login_required
def dashboard_view(request):
    """Vista del dashboard: prepara datos embebidos para las gráficas (sin API)."""
    from collections import Counter
    import json

    now = timezone.now()

    # KPIs básicos
    nuevos_pedidos = Pedido.objects.filter(fecha__gte=(now - datetime.timedelta(days=7))).count()
    total_usuarios = Usuario.objects.count()

    # Pre-cargar productos para mapear nombres/códigos -> categoría
    productos_qs = Producto.objects.all()
    prod_by_name = { (p.nombre_producto or '').strip().lower(): p for p in productos_qs }
    prod_by_code = { (p.codigo_producto or '').strip().lower(): p for p in productos_qs if p.codigo_producto }

    # Mapa valor->display de categoría
    # Obtener todas las categorías desde la tabla Categoria
    categorias_obj = Categoria.objects.all().order_by('nombre')
    categoria_map = {str(c.id): c.nombre for c in categorias_obj}

    # Inicializar conteos por categoría (usar nombres de la tabla Categoria)
    category_counts = { c.nombre: 0 for c in categorias_obj }

    # Contador por producto (nombre)
    product_counts = Counter()
    total_ventas = 0.0

    # Ventas por día para "nuevos usuarios" o para seguimiento (últimos 30 días)
    DAYS = 30
    start_date = (now - datetime.timedelta(days=DAYS - 1)).date()
    users_by_day = { (start_date + datetime.timedelta(days=i)): 0 for i in range(DAYS) }

    # Si el modelo Usuario tiene fecha_registro, agregamos nuevos usuarios por día
    has_user_date = hasattr(Usuario, 'fecha_registro')

    # Pedidos (puedes limitar este queryset por fecha si la BD es grande)
    pedidos_qs = Pedido.objects.select_related('usuario').order_by('-fecha')

    for ped in pedidos_qs:
        order_total = 0.0
        for item in (ped.productos or []):
            if isinstance(item, dict):
                # obtener price y qty de forma tolerante
                try:
                    price = float(item.get('precio', item.get('price', 0)) or 0)
                except Exception:
                    price = 0.0
                try:
                    qty = int(item.get('cantidad', item.get('qty', 1)) or 1)
                except Exception:
                    qty = 1

                order_total += price * qty

                # nombre/código para buscar el producto real y su categoría
                name = (item.get('nombre') or item.get('name') or '').strip()
                code = (item.get('codigo') or item.get('sku') or '').strip()

                prod = None
                if name:
                    prod = prod_by_name.get(name.lower())
                if prod is None and code:
                    prod = prod_by_code.get(code.lower())

                if prod is not None:
                    # si el producto tiene FK categoria, usar su nombre; si no, intentar usar el campo antiguo
                    cat_obj = getattr(prod, 'categoria', None)
                    if cat_obj:
                        cat_display = cat_obj.nombre
                    else:
                        cat_display = getattr(prod, 'categoria_producto', None) or 'Sin categoría'
                    # sumar qty a la categoría
                    category_counts[cat_display] = category_counts.get(cat_display, 0) + qty
                    # contar por producto usando el nombre real (más fiable)
                    product_counts[getattr(prod, 'nombre_producto', name) or name or 'Producto'] += qty
                else:
                    # fallback cuando no se encuentra producto en la tabla: usar el nombre del item
                    product_counts[name or 'Producto'] += qty
                    category_counts['Sin categoría'] = category_counts.get('Sin categoría', 0) + qty
            else:
                # item no dict -> contar como 1 unidad de 'desconocido'
                product_counts[str(item)] += 1

        total_ventas += order_total

        # contar nuevos usuarios por día (si aplica)
        if has_user_date:
            pass  # lo calculamos más abajo desde Usuario; aquí mantenemos ventas por día si lo quieres

    # Preparar datos para la gráfica de categorias (ordenadas descendente para destacar la más vendida)
    # reconstruir conteos por categoría para el menú (usar todas las categorias existentes)
    categorias_conteo = [(c.nombre, category_counts.get(c.nombre, 0)) for c in categorias_obj]
    marcas = Marca.objects.all().order_by('nombre')
    # ordenar por cantidad descendente
    sorted_categories = sorted(categorias_conteo, key=lambda x: x[1], reverse=True)
    category_labels = [c for c, v in sorted_categories]
    category_data = [v for c, v in sorted_categories]

    # Preparar top productos (top N)
    TOP_N = 8
    top_products = product_counts.most_common(TOP_N)
    product_labels = [t[0] for t in top_products]
    product_data = [t[1] for t in top_products]

    # Nuevos usuarios por día (últimos 30 días) si Usuario tiene fecha_registro
    new_users_labels = []
    new_users_data = []
    if hasattr(Usuario, 'fecha_creacion'):
        # construir mapa fecha -> count
        users_map = { (start_date + datetime.timedelta(days=i)): 0 for i in range(DAYS) }
        qs_users = Usuario.objects.filter(fecha_creacion__gte=start_date).values_list('fecha_creacion', flat=True)
        for dt in qs_users:
            try:
                d = dt if isinstance(dt, datetime.date) and not isinstance(dt, datetime.datetime) else dt.date()
                if d in users_map:
                    users_map[d] += 1
            except Exception:
                pass
        sorted_days = sorted(users_map.keys())
        new_users_labels = [d.strftime('%d/%m') for d in sorted_days]
        new_users_data = [users_map[d] for d in sorted_days]
    else:
        # si no hay fecha de creación, devolver 30 días vacíos para no romper la plantilla
        sorted_days = [start_date + datetime.timedelta(days=i) for i in range(DAYS)]
        new_users_labels = [d.strftime('%d/%m') for d in sorted_days]
        new_users_data = [0 for _ in sorted_days]


    # Pipeline y pedidos recientes (mantener como antes)
    pipeline = {
        'Nuevo': Pedido.objects.filter(estado__iexact='Pendiente').count(),
        'Procesando': Pedido.objects.filter(estado__icontains='proc').count(),
        'Enviado': Pedido.objects.filter(estado__icontains='envi').count(),
        'Entregado': Pedido.objects.filter(estado__icontains='entreg').count(),
    }

    recent_orders = []
    for ped in pedidos_qs[:8]:
        order_total = 0.0
        for item in (ped.productos or []):
            if isinstance(item, dict):
                try:
                    price = float(item.get('precio', item.get('price', 0)) or 0)
                except Exception:
                    price = 0.0
                try:
                    qty = int(item.get('cantidad', item.get('qty', 1)) or 1)
                except Exception:
                    qty = 1
                order_total += price * qty
        usuario_nombre = ped.usuario.nombre if ped.usuario else 'Anon'
        try:
            fecha_str = ped.fecha.astimezone(timezone.get_current_timezone()).strftime('%d/%m %H:%M') \
                        if getattr(ped.fecha, 'tzinfo', None) else ped.fecha.strftime('%d/%m %H:%M')
        except Exception:
            fecha_str = str(ped.fecha)
        recent_orders.append({
            'id': ped.id,
            'usuario_nombre': usuario_nombre,
            'fecha': fecha_str,
            'estado': ped.estado,
            'total': round(order_total, 2),
        })

    context = {
        'nuevos_pedidos': nuevos_pedidos,
        'total_usuarios': total_usuarios,
        'total_ventas': round(total_ventas, 2),

        # categorías -> se usan para el primer gráfico (ventas por categoría)
        'labels_json': json.dumps(category_labels),
        'data_json': json.dumps(category_data),

        # productos top -> se usan para el doughnut (producto más vendido)
        'product_labels_json': json.dumps(product_labels),
        'product_data_json': json.dumps(product_data),

        # nuevos usuarios (serie temporal) -> se usan para la gráfica de línea
        'ventas_series_labels_json': json.dumps(new_users_labels),
        'ventas_series_data_json': json.dumps(new_users_data),

        'pipeline': pipeline,
        'recent_orders': recent_orders,
    }
    return render(request, 'dashboard.html', context)

@login_required
def pedidos_dashboard_view(request):
    """Listado de pedidos con filtros para el dashboard admin."""
    id_q = request.GET.get('id', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()
    correo_q = request.GET.get('correo', '').strip()
    estado_q = request.GET.get('estado', '').strip()

    pedidos = Pedido.objects.select_related('usuario').order_by('-fecha')

    if id_q.isdigit():
        pedidos = pedidos.filter(id=id_q)
    if nombre_q:
        pedidos = pedidos.filter(usuario__nombre__icontains=nombre_q)
    if correo_q:
        pedidos = pedidos.filter(usuario__correo__icontains=correo_q)
    if estado_q:
        pedidos = pedidos.filter(estado__iexact=estado_q)

    return render(request, 'dashboard_pedidos.html', {
        'pedidos': pedidos,
        'filtros': {
            'id': id_q,
            'nombre': nombre_q,
            'correo': correo_q,
            'estado': estado_q,
        }
    })

@require_POST
@login_required
def cambiar_estado_pedido_view(request, pedido_id):
    """Cambia el estado de un pedido."""
    nuevo_estado = request.POST.get('estado')
    pedido = get_object_or_404(Pedido, id=pedido_id)
    pedido.estado = nuevo_estado
    pedido.save()
    return redirect('dashboard_pedidos')


# --------- ADMIN - PRODUCTOS ---------
@login_required
def AgregarProducto_view(request):
    """Agregar o editar productos desde el panel admin."""
    producto_editar = Producto.objects.get(id=request.POST.get('producto_id')) if request.method == 'POST' and request.POST.get('producto_id') else None
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto_editar)
        if form.is_valid():
            producto = form.save()
            for img in request.FILES.getlist('imagenes'):
                ImagenProducto.objects.create(producto=producto, imagen=img)
            messages.success(request, "Producto actualizado exitosamente." if producto_editar else "Producto registrado exitosamente.", extra_tags='dashboard')
            return redirect('AgregarProducto')
        else:
            # marcar error y enviar los errores del form para mostrar al usuario
            errores = []
            for f, errs in form.errors.items():
                errores.append(f"{f}: {'; '.join(errs)}")
            error_text = ' | '.join(errores) if errores else 'Error al procesar el producto.'
            messages.error(request, error_text, extra_tags='dashboard')
    else:
        form = ProductoForm(instance=producto_editar)
    productos_qs = Producto.objects.all()
    # construir lista de productos agotados para notificaciones
    out_of_stock_list = []
    for p in productos_qs:
        try:
            qty = p.stock.cantidad
        except Exception:
            qty = None
        if qty is None or qty <= 0:
            out_of_stock_list.append({'id': p.id, 'nombre': p.nombre_producto or '', 'codigo': p.codigo_producto or ''})
    # serializar a JSON para inyección segura en plantilla
    out_of_stock_json = json.dumps(out_of_stock_list, ensure_ascii=False)
    return render(request, 'AgregarProducto.html', {'form': form, 'productos': productos_qs, 'producto_editar': producto_editar, 'out_of_stock_json': out_of_stock_json})

@login_required
def editar_producto_view(request, producto_id):
    """Editar producto individual."""
    if request.method == 'POST':
        request.POST = request.POST.copy()
        request.POST['producto_id'] = producto_id
        return AgregarProducto_view(request)
    producto = get_object_or_404(Producto, id=producto_id)
    form = ProductoForm(instance=producto)
    productos_qs = Producto.objects.all()
    out_of_stock_list = []
    for p in productos_qs:
        try:
            qty = p.stock.cantidad
        except Exception:
            qty = None
        if qty is None or qty <= 0:
            out_of_stock_list.append({'id': p.id, 'nombre': p.nombre_producto or '', 'codigo': p.codigo_producto or ''})
    out_of_stock_json = json.dumps(out_of_stock_list, ensure_ascii=False)
    return render(request, 'AgregarProducto.html', {
        'form': form,
        'productos': productos_qs,
        'producto_editar': producto,
        'out_of_stock_json': out_of_stock_json,
    })

@login_required
def eliminar_producto_view(request, producto_id):
    """Eliminar un producto."""
    get_object_or_404(Producto, id=producto_id).delete()
    messages.success(request, "Producto eliminado exitosamente.", extra_tags='dashboard')
    return redirect('AgregarProducto')

# --------- ADMIN - CATEGORÍAS ---------
@login_required
def AgregarCategoria_view(request):
    """Agregar o editar categoría."""
    categoria_editar = Categoria.objects.get(id=request.POST.get('categoria_id')) if request.method == 'POST' and request.POST.get('categoria_id') else None
    form = CategoriaForm(request.POST or None, instance=categoria_editar)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Categoría actualizada exitosamente." if categoria_editar else "Categoría registrada exitosamente.", extra_tags='dashboard')
        return redirect('AgregarCategoria')
    return render(request, 'AgregarCategoria.html', {
        'form': form,
        'categorias': Categoria.objects.all(),
        'categoria_editar': categoria_editar
    })

@login_required
def editar_categoria_view(request, categoria_id):
    """Redirige a editar una categoría (usa la misma vista de agregar)."""
    # Mostrar formulario con la categoría cargada
    categoria = get_object_or_404(Categoria, id=categoria_id)
    form = CategoriaForm(request.POST or None, instance=categoria)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Categoría actualizada exitosamente.', extra_tags='dashboard')
        return redirect('dashboard_categorias')
    return render(request, 'categoria_form.html', {'form': form, 'accion': 'Editar'})

@login_required
def eliminar_categoria_view(request, categoria_id):
    """Elimina una categoría."""
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente.', extra_tags='dashboard')
        return redirect('dashboard_categorias')
    return render(request, 'confirm_delete.html', {'objeto': categoria, 'tipo': 'Categoría'})


# Nuevas vistas más explícitas para CRUD de categorías (compatibles con plantillas)
@login_required
def categorias_list(request):
    """Lista y filtros de categorías para el dashboard."""
    id_q = request.GET.get('id', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()

    categorias_qs = Categoria.objects.all()
    if id_q.isdigit():
        categorias_qs = categorias_qs.filter(id=id_q)
    if nombre_q:
        categorias_qs = categorias_qs.filter(nombre__icontains=nombre_q)

    return render(request, 'dashboard_categorias.html', {
        'categorias': categorias_qs,
        'filtros': {
            'id': id_q,
            'nombre': nombre_q,
        }
    })


@login_required
def categoria_agregar(request):
    """Agregar nueva categoría (formulario)."""
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría registrada exitosamente.', extra_tags='dashboard')
            return redirect('dashboard_categorias')
    else:
        form = CategoriaForm()
    return render(request, 'categoria_form.html', {'form': form, 'accion': 'Agregar'})


@login_required
def categoria_editar(request, pk):
    """Editar categoría existente."""
    categoria = get_object_or_404(Categoria, pk=pk)
    form = CategoriaForm(request.POST or None, instance=categoria)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Categoría actualizada exitosamente.', extra_tags='dashboard')
        return redirect('dashboard_categorias')
    return render(request, 'categoria_form.html', {'form': form, 'accion': 'Editar'})


@login_required
def categoria_eliminar(request, pk):
    """Eliminar categoría con confirmación."""
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente.', extra_tags='dashboard')
        return redirect('dashboard_categorias')
    return render(request, 'confirm_delete.html', {'objeto': categoria, 'tipo': 'Categoría'})

# --------- ADMIN - USUARIOS ---------
@login_required
def usuarios_lista(request):
    """Lista y filtro de usuarios."""
    id_q = request.GET.get('id', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()
    correo_q = request.GET.get('correo', '').strip()
    rol_q = request.GET.get('rol', '').strip()

    usuarios = Usuario.objects.all()
    if id_q.isdigit():
        usuarios = usuarios.filter(id=id_q)
    if nombre_q:
        usuarios = usuarios.filter(nombre__icontains=nombre_q)
    if correo_q:
        usuarios = usuarios.filter(correo__icontains=correo_q)
    if rol_q:
        usuarios = usuarios.filter(rol__icontains=rol_q)

    return render(request, 'dashboard_usuarios.html', {
        'usuarios': usuarios,
        'filtros': {
            'id': id_q,
            'nombre': nombre_q,
            'correo': correo_q,
            'rol': rol_q,
        }
    })

@login_required
def usuario_agregar(request):
    """Agregar nuevo usuario (admin)."""
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if request.POST.get('contraseña') != request.POST.get('confirmar_contraseña'):
            form.add_error(None, "Las contraseñas no coinciden.")
        elif form.is_valid():
            usuario = form.save(commit=False)
            usuario.contraseña = make_password(request.POST.get('contraseña'))
            usuario.save()
            return redirect('dashboard_usuarios')
    else:
        form = UsuarioForm()
    return render(request, 'Agregar_usuarios.html', {'form': form, 'accion': 'Agregar'})

@login_required
def usuario_editar(request, pk):
    """Editar usuario existente."""
    usuario = get_object_or_404(Usuario, pk=pk)
    form = UsuarioForm(request.POST or None, instance=usuario)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('dashboard_usuarios')
    return render(request, 'usuario_form.html', {'form': form, 'accion': 'Editar'})

@login_required
def usuario_eliminar(request, pk):
    """Eliminar un usuario."""
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        usuario.delete()
        return redirect('dashboard_usuarios')
    return render(request, 'usuario_confirmar_eliminar.html', {'usuario': usuario})

# --------- ADMIN - MARCAS ---------

@login_required
def marcas_list(request):
    """Lista y filtro de marcas."""
    id_q = request.GET.get('id', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()
    proveedor_q = request.GET.get('proveedor', '').strip()

    marcas_qs = Marca.objects.all()
    if id_q.isdigit():
        marcas_qs = marcas_qs.filter(id=id_q)
    if nombre_q:
        marcas_qs = marcas_qs.filter(nombre__icontains=nombre_q)
    if proveedor_q:
        marcas_qs = marcas_qs.filter(proveedores__nombre__icontains=proveedor_q)

    return render(request, 'dashboard_marcas.html', {
        'marcas': marcas_qs.distinct(),
        'filtros': {
            'id': id_q,
            'nombre': nombre_q,
            'proveedor': proveedor_q,
        }
    })

@login_required
def marca_agregar(request):
    """Agregar una nueva marca."""
    form = MarcaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('marcas_list')
    return render(request, 'marca_form.html', {'form': form, 'accion': 'Agregar'})

@login_required
def marca_editar(request, pk):
    """Editar una marca."""
    marca = get_object_or_404(Marca, pk=pk)
    form = MarcaForm(request.POST or None, instance=marca)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('marcas_list')
    return render(request, 'marca_form.html', {'form': form, 'accion': 'Editar'})

@login_required
def marca_eliminar(request, pk):
    """Eliminar una marca."""
    marca = get_object_or_404(Marca, pk=pk)
    if request.method == 'POST':
        marca.delete()
        return redirect('marcas_list')
    return render(request, 'confirm_delete.html', {'objeto': marca, 'tipo': 'Marca'})

# --------- ADMIN - PROVEEDORES ---------

@login_required
def dashboard_proveedores(request):
    """Lista y filtro de proveedores."""
    id_q = request.GET.get('id', '').strip()
    nombre_q = request.GET.get('nombre', '').strip()
    correo_q = request.GET.get('correo', '').strip()
    cedula_q = request.GET.get('cedula', '').strip()
    telefono_q = request.GET.get('telefono', '').strip()

    proveedores = Proveedor.objects.all()
    if id_q.isdigit():
        proveedores = proveedores.filter(id=id_q)
    if nombre_q:
        proveedores = proveedores.filter(nombre__icontains=nombre_q)
    if correo_q:
        proveedores = proveedores.filter(correo__icontains=correo_q)
    if cedula_q:
        proveedores = proveedores.filter(cedula__icontains=cedula_q)
    if telefono_q:
        proveedores = proveedores.filter(telefono__icontains=telefono_q)

    return render(request, 'dashboard_proveedores.html', {
        'proveedores': proveedores,
        'filtros': {
            'id': id_q,
            'nombre': nombre_q,
            'correo': correo_q,
            'cedula': cedula_q,
            'telefono': telefono_q,
        }
    })

@login_required
def proveedor_agregar(request):
    """Agregar proveedor nuevo."""
    form = ProveedorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('dashboard_proveedores')
    return render(request, 'proveedor_form.html', {'form': form, 'titulo': 'Agregar Proveedor'})

@login_required
def proveedor_editar(request, pk):
    """Editar proveedor."""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    form = ProveedorForm(request.POST or None, instance=proveedor)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('dashboard_proveedores')
    return render(request, 'proveedor_form.html', {'form': form, 'titulo': 'Editar Proveedor'})

@login_required
def proveedor_eliminar(request, pk):
    """Eliminar proveedor."""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('dashboard_proveedores')
    return render(request, 'proveedor_form.html', {'form': None, 'proveedor': proveedor, 'titulo': 'Eliminar Proveedor'})

# --------- PAGO ---------

@login_required
def pago_view(request):
    """Vista de pago y procesamiento del pedido."""
    if request.method == 'POST':
        carrito_json = request.POST.get('carrito')
        if carrito_json:
            carrito = json.loads(carrito_json)
            usuario = get_object_or_404(Usuario, id=request.session['usuario_id'])
            Pedido.objects.create(usuario=usuario, productos=carrito, estado='Pendiente')
            return redirect('index')
        messages.error(request, 'No se recibió el carrito.')
    return render(request, 'pago.html')

# --------- BUSCADOR GENERAL ---------

def buscar_view(request):
    """Buscar productos por ID, código o nombre."""
    id_busqueda = request.GET.get('id_busqueda', '').strip()
    codigo_busqueda = request.GET.get('codigo_busqueda', '').strip()
    nombre_busqueda = request.GET.get('nombre_busqueda', '').strip()

    productos = Producto.objects.all()
    filtros = Q()

    if id_busqueda.isdigit():
        filtros &= Q(id=int(id_busqueda))
    if codigo_busqueda:
        filtros &= Q(codigo_producto__icontains=codigo_busqueda)
    if nombre_busqueda:
        filtros &= Q(nombre_producto__icontains=nombre_busqueda)

    if filtros:
        productos = productos.filter(filtros)

    # Asegurar que siempre se pase un formulario al template para que los campos
    # como nombre/código se rendericen correctamente incluso al usar la búsqueda.
    from .forms import ProductoForm
    form = ProductoForm()
    productos_qs = productos
    out_of_stock_list = []
    for p in productos_qs:
        try:
            qty = p.stock.cantidad
        except Exception:
            qty = None
        if qty is None or qty <= 0:
            out_of_stock_list.append({'id': p.id, 'nombre': p.nombre_producto or '', 'codigo': p.codigo_producto or ''})
    out_of_stock_json = json.dumps(out_of_stock_list, ensure_ascii=False)
    return render(request, 'AgregarProducto.html', {
        'form': form,
        'productos': productos,
        'producto_editar': None,
        'id_busqueda': id_busqueda,
        'codigo_busqueda': codigo_busqueda,
        'nombre_busqueda': nombre_busqueda,
        'out_of_stock_json': out_of_stock_json,
    })



def verificar_correo(request):
    correo = request.GET.get('correo', '').strip().lower()
    existe = Usuario.objects.filter(correo__iexact=correo).exists()
    return JsonResponse({'existe': existe})



def verificar_correo_proveedor(request):
    correo = request.GET.get('correo', '').strip().lower()
    existe = Proveedor.objects.filter(correo__iexact=correo).exists()
    return JsonResponse({'existe': existe})

def validar_codigo_unico(request):
    codigo = request.GET.get('codigo', '')
    existe = Producto.objects.filter(codigo_producto=codigo).exists()
    return JsonResponse({'existe': existe})


import io
import json
import datetime
from django.http import HttpResponse
from django.contrib.staticfiles import finders
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule

def beautify_sheet(ws, *, col_widths=None, currency_cols=None, date_cols=None, wrap_cols=None, table_name=None, cond_format_rules=None):
    """
    Aplica formato agradable a una hoja openpyxl.
    Parámetros (opcionales):
      - col_widths: dict {col_index(1-based): width}
      - currency_cols: [col_index,...] columnas a formatear como moneda
      - date_cols: [col_index,...] columnas con formato fecha/hora
      - wrap_cols: [col_index,...] columnas con wrap_text
      - table_name: nombre para convertir rango en Table
      - cond_format_rules: [(col_idx, operator, formula, fill_color_hex), ...]
    """
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for col_idx in range(1, ws.max_column + 1):
        width = (col_widths or {}).get(col_idx, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if currency_cols:
        for col_idx in currency_cols:
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col_idx)
                c.number_format = '"$"#,##0.00'

    if date_cols:
        for col_idx in date_cols:
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col_idx)
                c.number_format = 'DD/MM/YYYY HH:MM'

    if wrap_cols:
        for col_idx in wrap_cols:
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col_idx)
                c.alignment = Alignment(wrap_text=True, vertical='top')

    ws.freeze_panes = 'A2'

    if table_name:
        try:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
        except Exception:
            pass

    if cond_format_rules:
        for (col_idx, operator, formula, fill_color) in cond_format_rules:
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}2:{col_letter}{ws.max_row}"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            rule = CellIsRule(operator=operator, formula=[formula], fill=fill)
            ws.conditional_formatting.add(rng, rule)

def beautify_sheet_elegant(ws, *, col_widths=None, currency_cols=None, date_cols=None, wrap_cols=None, table_name=None, cond_format_rules=None, logo_path=None):
    """
    """
    from openpyxl.drawing.image import Image as XLImage

    header_bg = "000000"        # encabezado negro
    header_color = "FFFFFF"     # texto blanco en encabezado
    border_color = "000000"     # bordes negros
    data_bg = "FFFFFF"          # filas base en blanco
    alt_fill = "D9D9D9"         # filas alternas en gris medio
    total_font_color = "000000" # texto negro en totales


    inserted_logo_row = False
    if logo_path:
        try:
            ws.insert_rows(1)
            ws.row_dimensions[1].height = 48
            try:
                img = XLImage(logo_path)
                img.width = 140
                img.height = 50
                ws.add_image(img, "A1")
            except Exception:
                pass
            inserted_logo_row = True
        except Exception:
            inserted_logo_row = False

    header_row = 2 if inserted_logo_row else 1
    data_start = header_row + 1

    header_fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
    try:
        for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
            cell.font = Font(bold=True, color=header_color, size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    except StopIteration:
        pass

    thin = Side(border_style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for r in range(data_start, ws.max_row + 1):
        for c_idx in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c_idx).fill = PatternFill(start_color=data_bg, end_color=data_bg, fill_type="solid")

    try:
        for r in range(data_start, ws.max_row + 1):
            if (r - data_start) % 2 == 0:
                for c_idx in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c_idx).fill = PatternFill(start_color=alt_fill, end_color=alt_fill, fill_type="solid")
    except Exception:
        pass

    for col_idx in range(1, ws.max_column + 1):
        width = (col_widths or {}).get(col_idx)
        if width is None:
            hdr_val = ws.cell(row=header_row, column=col_idx).value or ""
            width = max(10, min(48, int(len(str(hdr_val)) * 1.6)))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if currency_cols:
        for col_idx in currency_cols:
            for r in range(data_start, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = '"$"#,##0.00'
    if date_cols:
        for col_idx in date_cols:
            for r in range(data_start, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = 'DD/MM/YYYY HH:MM'
    if wrap_cols:
        for col_idx in wrap_cols:
            for r in range(data_start, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')

    try:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    except Exception:
        pass

    try:
        ws.freeze_panes = f"A{data_start}"
    except Exception:
        ws.freeze_panes = 'A2'

    try:
        for col_idx in range(1, ws.max_column + 1):
            hdr = str(ws.cell(row=header_row, column=col_idx).value or "").lower()
            if hdr.startswith("total") or "total" in hdr:
                for r in range(data_start, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).font = Font(bold=True, color=total_font_color)
    except Exception:
        pass

    if table_name:
        try:
            ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
        except Exception:
            pass

    if cond_format_rules:
        for (col_idx, operator, formula, fill_color) in cond_format_rules:
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}{data_start}:{col_letter}{ws.max_row}"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            rule = CellIsRule(operator=operator, formula=[formula], fill=fill)
            ws.conditional_formatting.add(rng, rule)

@login_required
def exportar_modelo_excel(request, modelo):
    """
    Exporta a Excel por 'modelo' (usuarios, proveedores, productos, pedidos, marcas).
    URL: /dashboard/exportar/<modelo>/
    """
    if request.session.get('usuario_rol') != 'admin':
        return HttpResponse('No autorizado', status=403)

    modelo = modelo.lower()
    filename = f"{modelo}_export.xlsx"

    # intentar localizar logo en static/images/logo.jpg
    logo_path = None
    try:
        logo_path = finders.find('images/logo.jpg')
    except Exception:
        logo_path = None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = modelo.capitalize()

    # ===== Usuarios =====
    if modelo == 'usuarios':
        headers = ["id", "nombre", "correo", "rol"]
        ws.append(headers)
        for u in Usuario.objects.all():
            ws.append([u.id, u.nombre, u.correo, getattr(u, 'rol', '')])

        beautify_sheet_elegant(
            ws,
            col_widths={1:6, 2:22, 3:30, 4:12},
            table_name="UsuariosTable",
            wrap_cols=[],
            logo_path=logo_path
        )

    # ===== Proveedores =====
    elif modelo == 'proveedores':
        headers = ["id", "nombre", "correo", "telefono", "cedula", "fecha_registro", "activo"]
        ws.append(headers)
        for p in Proveedor.objects.all():
            ws.append([p.id, p.nombre, p.correo or '', p.telefono or '', p.cedula or '', str(p.fecha_registro), bool(p.activo)])

        beautify_sheet_elegant(
            ws,
            col_widths={1:6,2:22,3:28,4:16,5:14,6:16,7:10},
            table_name="ProveedoresTable",
            wrap_cols=[3],
            logo_path=logo_path
        )

    # ===== Productos =====
    elif modelo == 'productos':
        headers = ["id", "codigo_producto", "nombre_producto", "valor_producto", "cantidad_producto", "categoria", "marca", "proveedor"]
        ws.append(headers)
        productos = Producto.objects.select_related('marca', 'proveedor').all()
        for pr in productos:
            marca = pr.marca.nombre if pr.marca else ''
            proveedor = pr.proveedor.nombre if pr.proveedor else ''
            categoria_display = pr.categoria.nombre if getattr(pr, 'categoria', None) else (getattr(pr, 'categoria_producto', '') or '')
            # Obtener cantidad desde el modelo Stock cuando exista; si no, intentar el atributo legado 'cantidad_producto'
            try:
                cantidad_val = pr.stock.cantidad if getattr(pr, 'stock', None) is not None else (getattr(pr, 'cantidad_producto', '') or '')
            except Exception:
                cantidad_val = getattr(pr, 'cantidad_producto', '') or ''
            ws.append([pr.id, pr.codigo_producto, pr.nombre_producto, float(pr.valor_producto), cantidad_val, categoria_display, marca, proveedor])

        beautify_sheet_elegant(
            ws,
            col_widths={1:6,2:16,3:36,4:14,5:12,6:18,7:18,8:18},
            currency_cols=[4],
            table_name="ProductosTable",
            cond_format_rules=[(5, 'lessThan', '5', 'FFC7CE')],
            logo_path=logo_path
        )

    # ===== Pedidos =====
    elif modelo == 'pedidos':
        headers = ["id", "usuario_id", "usuario_nombre", "estado", "fecha", "total_pedido", "productos"]
        ws.append(headers)
        pedidos = Pedido.objects.select_related('usuario').all().order_by('-fecha')
        for ped in pedidos:
            # Construir productos en formato legible por línea
            total_p = 0
            productos_text = ''
            try:
                formatted_lines = []
                for item in ped.productos:
                    if isinstance(item, dict):
                        nombre = item.get('nombre') or item.get('name') or ''
                        try:
                            precio = float(item.get('precio', item.get('price', 0)) or 0)
                        except Exception:
                            precio = 0.0
                        try:
                            cantidad = int(item.get('cantidad', item.get('qty', 1)) or 1)
                        except Exception:
                            cantidad = 1

                        # quitar claves de imagen por si quedan
                        item_copy = dict(item)
                        for key in ('imagen', 'foto', 'image', 'img', 'image_url', 'imagen_url'):
                            item_copy.pop(key, None)

                        precio_int = int(round(precio))
                        precio_str = f"${precio_int:,}".replace(",", ".")  # '149900' -> '149.900'

                        formatted_lines.append(f"{nombre} x{cantidad} - {precio_str} COP")
                        total_p += precio * cantidad
                    else:
                        formatted_lines.append(str(item))
                productos_text = "\n".join(formatted_lines)
            except Exception:
                productos_text = str(ped.productos)

            # Normalizar fecha para Excel (sin tzinfo)
            fecha_val = ped.fecha
            try:
                if isinstance(fecha_val, datetime.datetime) and fecha_val.tzinfo is not None:
                    try:
                        fecha_val = fecha_val.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                    except Exception:
                        fecha_val = fecha_val.replace(tzinfo=None)
            except Exception:
                fecha_val = str(ped.fecha)

            ws.append([ped.id, ped.usuario.id if ped.usuario else None, ped.usuario.nombre if ped.usuario else '', ped.estado, fecha_val, float(total_p), productos_text])

        beautify_sheet_elegant(
            ws,
            col_widths={1:6,2:8,3:22,4:12,5:20,6:14,7:60},
            currency_cols=[6],
            date_cols=[5],
            wrap_cols=[7],
            table_name="PedidosTable",
            logo_path=logo_path
        )

    # ===== Marcas =====
    elif modelo == 'marcas':
        headers = ["id", "nombre", "proveedores"]
        ws.append(headers)
        marcas = Marca.objects.prefetch_related('proveedores').all()
        for m in marcas:
            proveedores_nombres = ", ".join([p.nombre for p in m.proveedores.all()])
            ws.append([m.id, m.nombre, proveedores_nombres])

        beautify_sheet_elegant(
            ws,
            col_widths={1:6,2:28,3:40},
            table_name="MarcasTable",
            wrap_cols=[3],
            logo_path=logo_path
        )

    else:
        return HttpResponse('Modelo no soportado', status=400)

    # Guardar en buffer y devolver
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response

from .models import Blog
from .forms import BlogForm

# Listar todos los blogs
def blog_listado(request):
    blogs = Blog.objects.all()
    return render(request, 'blogs/listado.html', {'blogs': blogs})

# Agregar blog
def blog_agregar(request):
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('blog_listado')
    else:
        form = BlogForm()
    return render(request, 'blogs/formulario.html', {'form': form, 'accion': 'Agregar'})

# Editar blog
def blog_editar(request, id):
    blog = get_object_or_404(Blog, pk=id)
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            form.save()
            return redirect('blog_listado')
    else:
        form = BlogForm(instance=blog)
    return render(request, 'blogs/formulario.html', {'form': form, 'accion': 'Editar'})

# Eliminar blog
def blog_eliminar(request, id):
    blog = get_object_or_404(Blog, pk=id)
    if request.method == 'POST':
        blog.delete()
        return redirect('blog_listado')
    return render(request, 'blogs/confirmar_eliminar.html', {'blog': blog})

def dashboard_blogs(request):
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('blog_list')  # Redirige a sí misma para limpiar formulario
    else:
        form = BlogForm()

    blogs = Blog.objects.all()
    return render(request, 'dashboard_blog.html', {'form': form, 'blogs': blogs})
    

#Informe ventas PDF
def generar_informe_ventas(request):
    from .models import Pedido
    from django.http import HttpResponse
    try:
        import importlib
        pdfgen = importlib.import_module('reportlab.pdfgen')
        canvas = getattr(pdfgen, 'canvas', None)
        pagesizes = importlib.import_module('reportlab.lib.pagesizes')
        letter = getattr(pagesizes, 'letter', None)
        units = importlib.import_module('reportlab.lib.units')
        inch = getattr(units, 'inch', None)
        utils = importlib.import_module('reportlab.lib.utils')
        ImageReader = getattr(utils, 'ImageReader', None)
    except Exception:
        # reportlab no disponible en el entorno; las variables quedan en None
        canvas = None
        letter = None
        inch = None
        ImageReader = None
    # Si reportlab no está presente, devolver una respuesta clara en vez de fallar con AttributeError
    if canvas is None or letter is None or inch is None:
        msg = (
            "ReportLab no está disponible en este entorno. Para habilitar la generación de PDF instala la dependencia:\n"
            "pip install reportlab\n"
            "Luego reinicia el servidor de desarrollo y vuelve a intentar generar el informe."
        )
        return HttpResponse(msg, content_type='text/plain', status=500)
    from django.contrib.staticfiles import finders

    # Filtrar pedidos enviados
    pedidos = Pedido.objects.filter(estado="enviado")

    # Calcular total de ventas (suponiendo que tu modelo Pedido tiene propiedad .total)
    total_ventas = sum(p.total for p in pedidos)

    # Crear respuesta con PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="informe_ventas.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Buscar logo en static/images/logo.jpg
    logo_path = finders.find("images/logo.jpg")
    if logo_path:
        p.drawImage(ImageReader(logo_path), inch, height - 1.5*inch, width=150, height=60)

    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, height - inch, "Informe de Ventas - YSJI Fashi")

    # Texto descriptivo profesional
    p.setFont("Helvetica", 12)
    text = p.beginText(1*inch, height - 2.5*inch)
    text.setLeading(18)
    text.textLine("Este informe presenta un resumen actualizado del desempeño comercial de nuestra tienda de ropa.")
    text.textLine("Se consolidan únicamente los pedidos confirmados con estado 'Enviado', reflejando la actividad")
    text.textLine("operativa más reciente y los resultados obtenidos en términos de gestión de ventas.")
    text.textLine("")
    text.textLine(f"Pedidos completados y enviados: {pedidos.count()}")
    text.textLine(f"Valor total acumulado en ventas: ${total_ventas:,.0f} COP")
    p.drawText(text)

    # Finalizar
    p.showPage()
    p.save()
    return response

def eliminar_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)
    blog.delete()
    return redirect('blog_list')  # Aquí también usa el nombre correcto

def editar_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)
    blogs = Blog.objects.all()

    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            form.save()
            return redirect('blog_list')  # Nombre correcto de la vista principal
    else:
        form = BlogForm(instance=blog)

    return render(request, 'dashboard_blog.html', {'form': form, 'blogs': blogs, 'editando': True, 'blog_editado': blog})

def dashboard_blog(request):
    blogs = Blog.objects.all()
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('blog_list')  # Aquí también
    else:
        form = BlogForm()

    return render(request, 'dashboard_blog.html', {'form': form, 'blogs': blogs, 'editando': False})

# --------- CONTACTO ---------
from django.shortcuts import render
from .forms import MensajeForm

def Contactenos_view(request):
    print(">>> Vista Contactenos_view activada")  # Verifica que la vista se llama
    enviado = False

    if request.method == 'POST':
        print(">>> Método POST recibido")
        form = MensajeForm(request.POST)
        if form.is_valid():
            form.save()
            print(">>> Formulario válido y guardado")
            enviado = True
            form = MensajeForm()
        else:
            print(">>> Errores en el formulario:")
            print(form.errors)
    else:
        form = MensajeForm()

    return render(request, 'Contactenos.html', {'form': form, 'enviado': enviado})
from django.shortcuts import render
from .models import Mensaje

def dashboard_contacto(request):
    mensajes = Mensaje.objects.all().order_by('-fecha')
    return render(request, 'dashboard_contacto.html', {'mensajes': mensajes})

def eliminar_mensaje(request, mensaje_id):
    mensaje = get_object_or_404(Mensaje, id=mensaje_id)
    mensaje.delete()
    return redirect('dashboard_contacto') 

@require_POST
def eliminar_todos_mensajes(request):
    Mensaje.objects.all().delete()
    return redirect('dashboard_contacto')  